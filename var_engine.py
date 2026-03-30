# ================================================================
# var_engine.py - Moteur de calcul VaR CommandoQuant
# ================================================================
# OBJECTIF : Calculer la Value-at-Risk du portefeuille d'options
#            a partir des positions stockees dans SQL Server (tbl_Greeks)
#
# SOURCE DE DONNEES :
#   Serveur SQL Server : PERSO-AJE-DELL\BFASERVER01
#   Base de données    : CommandoQuant
#   Tables utilisées   : tbl_Greeks (positions avec Greeks calculés)
#
# 3 METHODES DE CALCUL :
#   1. VaR Parametrique  - formule analytique via Delta (rapide)
#   2. VaR Monte Carlo   - simulation de 10 000 scenarios (précis)
#   3. VaR Historique    - simulation basee sur des chocs reels (réaliste)
#
# ACCES DECOMMISSIONNE : Microsoft Access n'est plus utilisé.
#   Toutes les données transitent par SQL Server uniquement.
#
# OUTPUT : fichier var_results.xlsx dans le même dossier que ce script
#
# CONSEIL JUNIOR (le data analyst senior te parle) :
#   Tu te demandes pourquoi SQL Server et pas un fichier Excel ou Access ?
#   Imagine que 5 traders saisissent des positions en même temps.
#   Un fichier Excel, ça plante. Access, ça se corrompt.
#   SQL Server, c'est un vrai coffre-fort : transactions, verrous, audit.
#   En salle de marché, la donnée c'est de l'argent. On ne rigole pas.
# ================================================================

import math
import random
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------------------------------------------
# CONFIGURATION — SQL SERVER (PERSO-AJE-DELL\BFASERVER01)
# ----------------------------------------------------------------
# Authentification Windows (Trusted_Connection) : aucun mot de passe
# à stocker dans le code. Le compte Windows de l'utilisateur est
# utilisé pour se connecter. C'est la bonne pratique en entreprise.
SQL_SERVER  = r"PERSO-AJE-DELL\BFASERVER01"
SQL_DB      = "CommandoQuant"
OUTPUT_PATH = r"D:\ATELIER_IT\07 POC- sale de marché - commando\var_results.xlsx"

CONFIDENCE  = 0.95      # niveau de confiance 95%
N_SIMUL     = 10000     # nombre de simulations Monte Carlo
VOL_DAILY   = 0.015     # volatilite journaliere ~ vol annuelle / sqrt(252)
                        # 0.25 / sqrt(252) = 0.0157

# ----------------------------------------------------------------
# ETAPE 1 : LIRE LES POSITIONS DEPUIS SQL SERVER
# ----------------------------------------------------------------
def get_sql_connection():
    """
    Crée une connexion pyodbc vers SQL Server avec authentification Windows.

    CONSEIL JUNIOR (le data analyst senior t'explique comme si t'avais 5 ans) :
      Une "connexion" c'est comme décrocher le téléphone pour parler à SQL Server.
      "Trusted_Connection=yes" veut dire : je me connecte avec mon compte Windows,
      pas besoin de taper un mot de passe dans le code. C'est plus sécurisé.
      On essaie d'abord le driver "ODBC Driver 17" (le plus récent).
      Si pas installé, on bascule sur "SQL Server" (driver natif Windows).
    """
    # Essai avec le driver ODBC 17 for SQL Server (recommandé)
    drivers_a_essayer = [
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 13 for SQL Server",
        "SQL Server",
    ]

    conn_str_template = (
        "DRIVER={{{driver}}};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DB};"
        "Trusted_Connection=yes;"
        "TrustServerCertificate=yes;"
    )

    for driver in drivers_a_essayer:
        try:
            conn_str = conn_str_template.format(driver=driver)
            conn = pyodbc.connect(conn_str, timeout=10)
            print(f"[OK] Connecté à SQL Server via driver : {driver}")
            return conn
        except pyodbc.Error:
            continue

    raise ConnectionError(
        f"[ERREUR] Impossible de se connecter à {SQL_SERVER}\\{SQL_DB}.\n"
        "Vérifier : \n"
        "  1. Le service SQL Server est démarré (services.msc)\n"
        "  2. Le nom du serveur est correct : PERSO-AJE-DELL\\BFASERVER01\n"
        "  3. Un driver ODBC SQL Server est installé\n"
        "  4. Votre compte Windows a accès à la base CommandoQuant"
    )


def lire_positions():
    """
    Lit toutes les positions de tbl_Greeks dans SQL Server.
    Retourne une liste de dictionnaires, une position par ligne.

    CONSEIL JUNIOR :
      On lit uniquement les positions où Spot > 0 ET Vol > 0.
      Pourquoi ? Parce que si le spot est à 0, notre formule divise par 0.
      Et une option sur un actif à prix zéro... ça n'existe pas en vrai !
      C'est une "garde-fou" pour éviter les plantages.

    Chaque position contient :
      - underlying : ticker (ex: "TTE.PA")
      - option_type: "CALL" ou "PUT"
      - strike     : prix d'exercice
      - spot       : prix du sous-jacent au moment du pricing
      - vol        : volatilite utilisee
      - prix       : prix Black-Scholes calcule
      - delta      : sensibilite au spot (le plus important pour la VaR)
      - gamma      : variation du delta (correction d'ordre 2)
      - vega       : sensibilite a la vol
      - notional   : 100 000 EUR par defaut
    """

    conn = get_sql_connection()
    cursor = conn.cursor()

    # Les tables tbl_Greeks existent déjà sur PERSO-AJE-DELL\BFASERVER01
    # On lit la dernière valorisation de chaque position (ORDER BY GreeksId DESC)
    cursor.execute("""
        SELECT TOP 1 WITH TIES
               g.Underlying, g.OptionType, g.Strike, g.Spot, g.ImpliedVol,
               g.Price, g.Delta, g.Gamma, g.Vega
        FROM   tbl_Greeks g
        WHERE  g.Spot > 0
          AND  g.ImpliedVol > 0
        ORDER BY ROW_NUMBER() OVER (
            PARTITION BY g.Underlying, g.OptionType, g.Strike
            ORDER BY g.CalcDate DESC, g.GreeksId DESC
        )
    """)

    positions = []
    for row in cursor.fetchall():
        positions.append({
            "underlying" : row[0],
            "option_type": row[1],
            "strike"     : float(row[2]) if row[2] else 0,
            "spot"       : float(row[3]) if row[3] else 0,
            "vol"        : float(row[4]) if row[4] else 0,
            "prix"       : float(row[5]) if row[5] else 0,
            "delta"      : float(row[6]) if row[6] else 0,
            "gamma"      : float(row[7]) if row[7] else 0,
            "vega"       : float(row[8]) if row[8] else 0,
            "notional"   : 100000
        })

    conn.close()
    print(f"[OK] {len(positions)} positions lues depuis SQL Server ({SQL_SERVER})")
    return positions


# ----------------------------------------------------------------
# ETAPE 2 : VaR PARAMETRIQUE (methode Delta)
# ----------------------------------------------------------------
def var_parametrique(positions, confidence=0.95):
    """
    Methode la plus simple et la plus rapide.

    FORMULE :
      dP = Delta x dS   (approximation 1er ordre)
      dS = spot x vol_journaliere x z
      z  = quantile loi normale = 1.6449 pour 95%

    HYPOTHESE : variations de spot suivent une loi normale.
    LIMITE    : ignore la convexite (Gamma).

    RETOURNE : VaR en euros (perte maximale a 95%)
    """

    # Quantile de la loi normale a 95%
    # P(X < -1.6449) = 5%
    z_95 = 1.6449

    variance_totale = 0

    for p in positions:
        # Nombre d'actions = Notional / Spot
        nb_actions = p["notional"] / p["spot"] if p["spot"] > 0 else 0

        # Variation du spot en euros (1 sigma journalier)
        sigma_spot = p["spot"] * VOL_DAILY

        # Sensibilite du portefeuille via Delta
        # dP = Delta x nb_actions x dS
        sensibilite = p["delta"] * nb_actions * sigma_spot

        # Contribution a la variance totale
        variance_totale += sensibilite ** 2

    # Ecart-type du portefeuille
    sigma_portfolio = math.sqrt(variance_totale)

    # VaR = z x sigma
    var = z_95 * sigma_portfolio

    print(f"[VaR Parametrique 95%] = {var:.2f} EUR")
    return var


# ----------------------------------------------------------------
# ETAPE 3 : VaR MONTE CARLO
# ----------------------------------------------------------------
def var_monte_carlo(positions, n_simul=10000, confidence=0.95):
    """
    Simule N scenarios de variation du marche aleatoirement.

    FORMULE par simulation i :
      dS_i = spot x vol_journaliere x epsilon_i  (epsilon ~ N(0,1))
      dP_i = Delta x dS_i + 0.5 x Gamma x dS_i^2

    Le terme Gamma (ordre 2) capture la convexite.
    Plus precis que la methode parametrique.

    RETOURNE : VaR en euros + liste des P&L simules
    """

    pnl_simulations = []

    for i in range(n_simul):
        pnl_scenario = 0

        for p in positions:
            if p["spot"] <= 0:
                continue

            nb_actions = p["notional"] / p["spot"]

            # Tirage aleatoire normal via Box-Muller
            # epsilon suit une loi normale N(0,1)
            u1 = random.random()
            u2 = random.random()
            epsilon = math.sqrt(-2 * math.log(u1)) * math.cos(2 * math.pi * u2)

            # Variation du spot
            dS = p["spot"] * VOL_DAILY * epsilon

            # P&L via Taylor ordre 2 (Delta + Gamma)
            dP = (p["delta"] * dS + 0.5 * p["gamma"] * dS ** 2) * nb_actions
            pnl_scenario += dP

        pnl_simulations.append(pnl_scenario)

    # Trier du plus mauvais au meilleur
    pnl_simulations.sort()

    # Quantile a 5% = VaR 95%
    idx_var = int((1 - confidence) * n_simul)
    var = -pnl_simulations[idx_var]

    print(f"[VaR Monte Carlo 95%] = {var:.2f} EUR ({n_simul} simulations)")
    return var, pnl_simulations


# ----------------------------------------------------------------
# ETAPE 4 : VaR HISTORIQUE (scenarios de stress)
# ----------------------------------------------------------------
def var_historique(positions, confidence=0.95):
    """
    Rejoue des chocs de marche historiques reels.

    SCENARIOS : variations journalieres observees lors de vraies crises
      - COVID mars 2020      : -12%
      - Lehman sept 2008     : -9.5%
      - Flash Crash 2010     : -7.2%
      - Crise dettes 2011    : -6.8%
      - Brexit 2016          : -5.5%
      - Guerre Ukraine 2022  : -4.8%

    RETOURNE : VaR en euros + liste des scenarios
    """

    chocs_historiques = [
        -0.120,   # COVID mars 2020
        -0.095,   # Lehman Brothers sept 2008
        -0.072,   # Flash Crash mai 2010
        -0.068,   # Crise dettes souveraines 2011
        -0.055,   # Brexit juin 2016
        -0.048,   # Guerre Ukraine fev 2022
        -0.042,   # Crise bancaire mars 2023
        -0.035,
        -0.028,
        -0.022,
        -0.018,
        -0.015,
        -0.012,
        -0.010,
        -0.008,
        -0.006,
        0.005,
        0.010,
        0.015,
        0.020,
        0.030,
        0.045,
        0.060,
        0.075,
        0.085,   # rebond post-COVID
    ]

    pnl_historique = []

    for choc in chocs_historiques:
        pnl_scenario = 0

        for p in positions:
            if p["spot"] <= 0:
                continue

            nb_actions = p["notional"] / p["spot"]
            dS = p["spot"] * choc
            dP = (p["delta"] * dS + 0.5 * p["gamma"] * dS ** 2) * nb_actions
            pnl_scenario += dP

        pnl_historique.append((choc * 100, pnl_scenario))

    # Trier par P&L croissant
    pnl_historique.sort(key=lambda x: x[1])

    idx_var = int((1 - confidence) * len(pnl_historique))
    var = -pnl_historique[idx_var][1]

    print(f"[VaR Historique 95%] = {var:.2f} EUR")
    return var, pnl_historique


# ----------------------------------------------------------------
# ETAPE 5 : EXPORTER LES RESULTATS DANS EXCEL
# ----------------------------------------------------------------
def exporter_excel(positions, var_param, var_mc, pnl_mc,
                   var_histo, pnl_histo):
    """
    Cree var_results.xlsx avec 3 feuilles :
      - VaR Summary      : les 3 methodes + positions
      - Monte Carlo      : distribution des P&L simules
      - Historique       : scenarios de stress
    """

    wb = openpyxl.Workbook()

    # Couleurs CommandoQuant
    DARK  = "0B1426"
    AMBER = "D4880A"
    GOLD  = "F0B429"
    GREEN = "00C853"
    RED   = "E53935"
    NAVY  = "0D2040"
    WHITE = "FFFFFF"
    NAVY2 = "112855"

    def style_header(cell, bg=NAVY, fg=GOLD):
        cell.font = Font(bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_title(cell):
        cell.font = Font(bold=True, color=GOLD, size=14)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================
    # FEUILLE 1 : VAR SUMMARY
    # ============================================================
    ws1 = wb.active
    ws1.title = "VaR Summary"
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A1:F1")
    style_title(ws1["A1"])
    ws1["A1"] = "CommandoQuant - Value-at-Risk Report"

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Portefeuille : {len(positions)} positions | Confiance : 95% | Horizon : 1 jour"
    ws1["A2"].font = Font(italic=True, color="B0BEC5", size=9)
    ws1["A2"].fill = PatternFill("solid", fgColor=DARK)
    ws1["A2"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Methode", "VaR 95% (EUR)", "Interpretation", "Fiabilite"], 1):
        style_header(ws1.cell(4, col, h))

    data_var = [
        ("Parametrique (Delta)",      round(var_param, 2), "Formule analytique via sensibilite Delta",     "Rapide - ignore Gamma"),
        ("Monte Carlo (10 000 simul)",round(var_mc,    2), "Simulation 10 000 scenarios aleatoires",       "Precis - inclut Delta + Gamma"),
        ("Historique (crises reelles)",round(var_histo, 2),"Base sur COVID, Lehman, Flash Crash...",       "Realiste - limite aux crises connues"),
    ]

    for i, (methode, var_val, interp, fiab) in enumerate(data_var):
        row = 5 + i
        ws1.cell(row, 1, methode)
        cell_var = ws1.cell(row, 2, var_val)
        cell_var.font = Font(bold=True, color=RED, size=11)
        ws1.cell(row, 3, interp)
        ws1.cell(row, 4, fiab)

        bg = NAVY if i % 2 == 0 else NAVY2
        for col in range(1, 5):
            ws1.cell(row, col).fill = PatternFill("solid", fgColor=bg)
            if col != 2:
                ws1.cell(row, col).font = Font(color=WHITE)

    # Section positions
    ws1.merge_cells("A9:F9")
    style_header(ws1["A9"], bg=AMBER, fg=DARK)
    ws1["A9"] = "POSITIONS DU PORTEFEUILLE"

    for col, h in enumerate(["Underlying", "Type", "Strike", "Spot", "Delta", "Contrib VaR (EUR)"], 1):
        style_header(ws1.cell(10, col, h))

    for i, p in enumerate(positions):
        row = 11 + i
        nb = p["notional"] / p["spot"] if p["spot"] > 0 else 0
        contrib = abs(p["delta"] * nb * p["spot"] * VOL_DAILY * 1.6449)

        ws1.cell(row, 1, p["underlying"])
        ws1.cell(row, 2, p["option_type"])
        ws1.cell(row, 3, round(p["strike"], 2))
        ws1.cell(row, 4, round(p["spot"], 2))
        ws1.cell(row, 5, round(p["delta"], 4))
        ws1.cell(row, 6, round(contrib, 2))

        bg = NAVY if i % 2 == 0 else NAVY2
        for col in range(1, 7):
            ws1.cell(row, col).fill = PatternFill("solid", fgColor=bg)
            ws1.cell(row, col).font = Font(color=WHITE)

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 45
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 18

    # ============================================================
    # FEUILLE 2 : MONTE CARLO
    # ============================================================
    ws2 = wb.create_sheet("Monte Carlo")
    ws2.merge_cells("A1:B1")
    style_title(ws2["A1"])
    ws2["A1"] = "Distribution P&L - Monte Carlo (200 premiers)"

    for col, h in enumerate(["Scenario #", "P&L (EUR)"], 1):
        style_header(ws2.cell(3, col, h))

    for i, pnl in enumerate(pnl_mc[:200]):
        row = 4 + i
        ws2.cell(row, 1, i + 1)
        cell = ws2.cell(row, 2, round(pnl, 2))
        cell.font = Font(color=RED if pnl < 0 else GREEN, bold=True)
        bg = NAVY if i % 2 == 0 else NAVY2
        ws2.cell(row, 1).fill = PatternFill("solid", fgColor=bg)
        ws2.cell(row, 2).fill = PatternFill("solid", fgColor=bg)
        ws2.cell(row, 1).font = Font(color=WHITE)

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 15

    # ============================================================
    # FEUILLE 3 : HISTORIQUE
    # ============================================================
    ws3 = wb.create_sheet("Historique")
    ws3.merge_cells("A1:C1")
    style_title(ws3["A1"])
    ws3["A1"] = "Scenarios Historiques - Stress Tests"

    for col, h in enumerate(["Choc Spot (%)", "P&L Portefeuille (EUR)", "Contexte"], 1):
        style_header(ws3.cell(3, col, h))

    interpretations = {
        -12.0: "COVID mars 2020",
        -9.5 : "Lehman Brothers 2008",
        -7.2 : "Flash Crash 2010",
        -6.8 : "Crise dettes 2011",
        -5.5 : "Brexit 2016",
        -4.8 : "Guerre Ukraine 2022",
        -4.2 : "Crise bancaire 2023",
    }

    for i, (choc_pct, pnl) in enumerate(pnl_histo):
        row = 4 + i
        label = interpretations.get(round(choc_pct, 1),
                "Hausse marche" if choc_pct > 0 else "Baisse normale")

        ws3.cell(row, 1, f"{choc_pct:+.1f}%")
        cell = ws3.cell(row, 2, round(pnl, 2))
        cell.font = Font(color=RED if pnl < 0 else GREEN, bold=True)
        ws3.cell(row, 3, label)

        bg = NAVY if i % 2 == 0 else NAVY2
        for col in range(1, 4):
            ws3.cell(row, col).fill = PatternFill("solid", fgColor=bg)
            if col != 2:
                ws3.cell(row, col).font = Font(color=WHITE)

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 25

    wb.save(OUTPUT_PATH)
    print(f"[OK] Resultats exportes : {OUTPUT_PATH}")
    print(f"     Source données : SQL Server {SQL_SERVER} / {SQL_DB}")


# ================================================================
# PROGRAMME PRINCIPAL
# ================================================================
if __name__ == "__main__":

    print("=" * 50)
    print("CommandoQuant - Calcul VaR")
    print("=" * 50)

    positions = lire_positions()

    if not positions:
        print("[ERREUR] Aucune position trouvee dans la base.")
        exit(1)

    print(f"\n{len(positions)} positions chargees :")
    for p in positions:
        print(f"  {p['underlying']} {p['option_type']} "
              f"K={p['strike']} S={p['spot']} Delta={p['delta']:.4f}")

    print("\n--- Calcul VaR ---")
    var_param            = var_parametrique(positions)
    var_mc, pnl_mc       = var_monte_carlo(positions, N_SIMUL)
    var_histo, pnl_histo = var_historique(positions)

    print("\n--- Export Excel ---")
    exporter_excel(positions, var_param, var_mc, pnl_mc,
                   var_histo, pnl_histo)

    print("\n" + "=" * 50)
    print("VaR RESUME")
    print("=" * 50)
    print(f"Parametrique : {var_param:>10.2f} EUR")
    print(f"Monte Carlo  : {var_mc:>10.2f} EUR")
    print(f"Historique   : {var_histo:>10.2f} EUR")
    print("=" * 50)
    print("Termine.")
