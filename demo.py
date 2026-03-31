#!/usr/bin/env python3
# ================================================================
# demo.py - Démo autonome CommandoQuant
# ================================================================
# Lance le pipeline VaR complet SANS SQL Server.
# Utilise les positions exemples depuis data/sample_positions.csv.
#
# Utilisation :
#   python demo.py
#   python demo.py --output results/mon_rapport_var.xlsx
#
# Prérequis : pip install openpyxl
# Optionnel  : pip install pandas  (pour un chargement CSV enrichi)
# ================================================================

import csv
import math
import os
import sys
import argparse
import random


# ----------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------
DEFAULT_INPUT  = os.path.join(os.path.dirname(__file__), "data", "sample_positions.csv")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "var_results_demo.xlsx")

CONFIDENCE = 0.95
N_SIMUL    = 10000
VOL_DAILY  = 0.015    # ~25% annualisée / sqrt(252)


# ----------------------------------------------------------------
# 1. Chargement des positions depuis CSV (sans SQL Server)
# ----------------------------------------------------------------
def load_positions(csv_path: str) -> list:
    """
    Charge les positions d'options depuis un fichier CSV.

    Colonnes attendues :
      underlying, option_type, strike, spot, vol, prix,
      delta, gamma, vega, notional

    Retourne une liste de dictionnaires — même structure que lire_positions()
    dans var_engine.py (mode SQL).
    """
    positions = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            positions.append({
                "underlying":  row["underlying"],
                "option_type": row["option_type"],
                "strike":      float(row["strike"]),
                "spot":        float(row["spot"]),
                "vol":         float(row["vol"]),
                "prix":        float(row["prix"]),
                "delta":       float(row["delta"]),
                "gamma":       float(row["gamma"]),
                "vega":        float(row["vega"]),
                "notional":    float(row["notional"]),
            })

    print(f"[OK] {len(positions)} positions chargees depuis {csv_path}")
    return positions


# ----------------------------------------------------------------
# 2. VaR Paramétrique (méthode Delta)
# ----------------------------------------------------------------
def var_parametrique(positions: list, confidence: float = 0.95) -> float:
    """
    VaR Paramétrique — approximation de Taylor au 1er ordre via Delta.

    Formule :  VaR = z * sqrt( sum( (Delta * nb * sigma_spot)^2 ) )
    z(95%)  = 1.6449

    Rapide mais ignore la convexité (Gamma).
    """
    z = 1.6449
    variance = 0.0

    for p in positions:
        if p["spot"] <= 0:
            continue
        # Nombre d'actions équivalent = Notional / Spot
        nb         = p["notional"] / p["spot"]
        # Variation du spot en euros (1 sigma journalier)
        sigma_spot = p["spot"] * VOL_DAILY
        # Contribution via Delta
        contrib    = p["delta"] * nb * sigma_spot
        variance  += contrib ** 2

    var = z * math.sqrt(variance)
    print(f"  [VaR Parametrique 95%]   {var:>12,.2f} EUR")
    return var


# ----------------------------------------------------------------
# 3. VaR Monte Carlo (Delta + Gamma)
# ----------------------------------------------------------------
def var_monte_carlo(positions: list, n_simul: int = 10000,
                    confidence: float = 0.95, seed: int = None):
    """
    VaR Monte Carlo — N scénarios de marché aléatoires.

    dP_i = Delta * dS_i + 0.5 * Gamma * dS_i^2
    dS_i = spot * vol_journaliere * epsilon_i  (epsilon ~ N(0,1))

    Plus précis que la méthode paramétrique : capture la convexité via Gamma.
    Retourne (var, liste_pnl).
    """
    if seed is not None:
        random.seed(seed)

    pnl_list = []

    for _ in range(n_simul):
        pnl = 0.0
        for p in positions:
            if p["spot"] <= 0:
                continue
            nb = p["notional"] / p["spot"]

            # Transformation Box-Muller → N(0,1)
            u1 = random.random() or 1e-10
            u2 = random.random()
            eps = math.sqrt(-2.0 * math.log(u1)) * math.cos(2.0 * math.pi * u2)

            # Variation du spot
            dS  = p["spot"] * VOL_DAILY * eps
            # P&L via Taylor ordre 2 (Delta + Gamma)
            dP  = (p["delta"] * dS + 0.5 * p["gamma"] * dS ** 2) * nb
            pnl += dP

        pnl_list.append(pnl)

    # Trier du plus mauvais au meilleur
    pnl_list.sort()
    # Quantile à 5% = VaR 95%
    idx = int((1.0 - confidence) * n_simul)
    var = -pnl_list[idx]

    print(f"  [VaR Monte Carlo  95%]  {var:>12,.2f} EUR  ({n_simul:,} simulations)")
    return var, pnl_list


# ----------------------------------------------------------------
# 4. VaR Historique (chocs de crises réelles)
# ----------------------------------------------------------------
def var_historique(positions: list, confidence: float = 0.95):
    """
    VaR Historique — rejoue des chocs de marché historiques réels.

    Scénarios : COVID-19, Lehman Brothers, Flash Crash,
    crise dettes européennes, Brexit, guerre Ukraine, crise SVB.

    Retourne (var, [(choc_pct, pnl, label), ...]).
    """
    shocks = [
        (-0.120, "COVID-19 mars 2020"),
        (-0.095, "Lehman Brothers sept. 2008"),
        (-0.072, "Flash Crash mai 2010"),
        (-0.068, "Crise dettes souveraines EU 2011"),
        (-0.055, "Brexit juin 2016"),
        (-0.048, "Russie-Ukraine fev. 2022"),
        (-0.042, "Crise SVB mars 2023"),
        (-0.035, "Stress -3.5%"),
        (-0.028, "Stress -2.8%"),
        (-0.022, "Stress -2.2%"),
        (-0.018, "Stress -1.8%"),
        (-0.015, "Normal -1.5%"),
        (-0.012, "Normal -1.2%"),
        (-0.010, "Normal -1.0%"),
        (-0.008, "Normal -0.8%"),
        (-0.006, "Normal -0.6%"),
        ( 0.005, "Rebond +0.5%"),
        ( 0.010, "Rebond +1.0%"),
        ( 0.015, "Rebond +1.5%"),
        ( 0.020, "Hausse +2.0%"),
        ( 0.030, "Hausse +3.0%"),
        ( 0.045, "Bull Run +4.5%"),
        ( 0.060, "Bull Run +6.0%"),
        ( 0.075, "Fort rallye +7.5%"),
        ( 0.085, "Rebond post-COVID"),
    ]

    results = []
    for choc, label in shocks:
        pnl = 0.0
        for p in positions:
            if p["spot"] <= 0:
                continue
            nb  = p["notional"] / p["spot"]
            dS  = p["spot"] * choc
            dP  = (p["delta"] * dS + 0.5 * p["gamma"] * dS ** 2) * nb
            pnl += dP
        results.append((choc * 100, pnl, label))

    # Trier par P&L croissant (pire en premier)
    results.sort(key=lambda x: x[1])
    idx = int((1.0 - confidence) * len(results))
    var = -results[idx][1]

    print(f"  [VaR Historique   95%]   {var:>12,.2f} EUR  ({len(shocks)} scenarios)")
    return var, results


# ----------------------------------------------------------------
# 5. Export Excel (rapport stylisé)
# ----------------------------------------------------------------
def export_excel(positions, var_param, var_mc, pnl_mc,
                 var_histo, pnl_histo, output_path: str):
    """
    Génère un rapport Excel stylisé avec 3 feuilles :
      - VaR Summary   : comparaison des 3 méthodes + détail des positions
      - Monte Carlo   : 200 premiers P&L simulés
      - Historique    : tous les résultats des stress tests de crise
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[ATTENTION] openpyxl non installe. Export Excel ignore.")
        print("            Installer avec : pip install openpyxl")
        return

    # Palette de couleurs CommandoQuant
    DARK  = "0B1426"
    AMBER = "D4880A"
    GOLD  = "F0B429"
    GREEN = "00C853"
    RED   = "E53935"
    NAVY  = "0D2040"
    NAVY2 = "112855"
    WHITE = "FFFFFF"

    def hdr(cell, bg=NAVY, fg=GOLD):
        cell.font      = Font(bold=True, color=fg, size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def title(cell):
        cell.font      = Font(bold=True, color=GOLD, size=14)
        cell.fill      = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # ---- Feuille 1 : Résumé VaR ----
    ws = wb.active
    ws.title = "VaR Summary"
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A1:F1")
    title(ws["A1"])
    ws["A1"] = "CommandoQuant  —  Rapport Value-at-Risk  (Mode Demo)"

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Portefeuille : {len(positions)} positions  |  "
        f"Confiance : {int(CONFIDENCE*100)}%  |  Horizon : 1 jour  |  "
        f"Vol : {VOL_DAILY*100:.2f}%/jour"
    )
    ws["A2"].font      = Font(italic=True, color="B0BEC5", size=9)
    ws["A2"].fill      = PatternFill("solid", fgColor=DARK)
    ws["A2"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Methode", "VaR 95% (EUR)", "Description", "Fiabilite"], 1):
        hdr(ws.cell(4, col, h))

    rows_var = [
        ("Parametrique (Delta)",        round(var_param, 2),
         "Formule analytique — sensibilite Delta uniquement",  "Rapide / ignore Gamma"),
        ("Monte Carlo (10 000 simul.)", round(var_mc,    2),
         f"10 000 scenarios de marche aleatoires (Delta+Gamma)", "Precis / plus lourd"),
        ("Historique (crises reelles)", round(var_histo, 2),
         "COVID, Lehman, Flash Crash, Brexit, Ukraine…",       "Realiste / crises connues"),
    ]

    for i, (method, val, desc, rel) in enumerate(rows_var):
        r = 5 + i
        ws.cell(r, 1, method)
        c = ws.cell(r, 2, val)
        c.font = Font(bold=True, color=RED, size=11)
        ws.cell(r, 3, desc)
        ws.cell(r, 4, rel)
        bg = NAVY if i % 2 == 0 else NAVY2
        for col in range(1, 5):
            ws.cell(r, col).fill = PatternFill("solid", fgColor=bg)
            if col != 2:
                ws.cell(r, col).font = Font(color=WHITE)

    ws.merge_cells("A9:F9")
    hdr(ws["A9"], bg=AMBER, fg=DARK)
    ws["A9"] = "POSITIONS DU PORTEFEUILLE"

    for col, h in enumerate(
        ["Sous-jacent", "Type", "Strike", "Spot", "Delta", "Contribution VaR (EUR)"], 1
    ):
        hdr(ws.cell(10, col, h))

    for i, p in enumerate(positions):
        r   = 11 + i
        nb  = p["notional"] / p["spot"] if p["spot"] > 0 else 0
        con = abs(p["delta"] * nb * p["spot"] * VOL_DAILY * 1.6449)
        ws.cell(r, 1, p["underlying"])
        ws.cell(r, 2, p["option_type"])
        ws.cell(r, 3, round(p["strike"],  2))
        ws.cell(r, 4, round(p["spot"],    2))
        ws.cell(r, 5, round(p["delta"],   4))
        ws.cell(r, 6, round(con,          2))
        bg = NAVY if i % 2 == 0 else NAVY2
        for col in range(1, 7):
            ws.cell(r, col).fill = PatternFill("solid", fgColor=bg)
            ws.cell(r, col).font = Font(color=WHITE)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 22

    # ---- Feuille 2 : Distribution Monte Carlo ----
    ws2 = wb.create_sheet("Monte Carlo")
    ws2.merge_cells("A1:B1")
    title(ws2["A1"])
    ws2["A1"] = "Distribution P&L — Monte Carlo (200 premiers scenarios)"
    for col, h in enumerate(["Scenario #", "P&L (EUR)"], 1):
        hdr(ws2.cell(3, col, h))
    for i, pnl in enumerate(pnl_mc[:200]):
        r = 4 + i
        ws2.cell(r, 1, i + 1)
        c = ws2.cell(r, 2, round(pnl, 2))
        c.font = Font(color=RED if pnl < 0 else GREEN, bold=True)
        bg = NAVY if i % 2 == 0 else NAVY2
        for col in (1, 2):
            ws2.cell(r, col).fill = PatternFill("solid", fgColor=bg)
        ws2.cell(r, 1).font = Font(color=WHITE)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16

    # ---- Feuille 3 : Stress tests historiques ----
    ws3 = wb.create_sheet("Historique")
    ws3.merge_cells("A1:C1")
    title(ws3["A1"])
    ws3["A1"] = "Stress Tests Historiques — Scenarios de Crises Reelles"
    for col, h in enumerate(["Choc Spot (%)", "P&L Portefeuille (EUR)", "Evenement"], 1):
        hdr(ws3.cell(3, col, h))
    for i, (choc, pnl, label) in enumerate(pnl_histo):
        r = 4 + i
        ws3.cell(r, 1, f"{choc:+.1f}%")
        c = ws3.cell(r, 2, round(pnl, 2))
        c.font = Font(color=RED if pnl < 0 else GREEN, bold=True)
        ws3.cell(r, 3, label)
        bg = NAVY if i % 2 == 0 else NAVY2
        for col in (1, 2, 3):
            ws3.cell(r, col).fill = PatternFill("solid", fgColor=bg)
            if col != 2:
                ws3.cell(r, col).font = Font(color=WHITE)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 30

    wb.save(output_path)
    print(f"\n[OK] Rapport genere : {output_path}")


# ----------------------------------------------------------------
# Point d'entrée principal
# ----------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="CommandoQuant — Demo VaR autonome (sans SQL Server)"
    )
    parser.add_argument(
        "--input", default=DEFAULT_INPUT,
        help=f"Fichier CSV des positions (defaut : {DEFAULT_INPUT})"
    )
    parser.add_argument(
        "--output", default=DEFAULT_OUTPUT,
        help=f"Chemin du rapport Excel de sortie (defaut : {DEFAULT_OUTPUT})"
    )
    parser.add_argument(
        "--simulations", type=int, default=N_SIMUL,
        help=f"Nombre de simulations Monte Carlo (defaut : {N_SIMUL})"
    )
    parser.add_argument(
        "--seed", type=int, default=42,
        help="Graine aleatoire pour la reproductibilite Monte Carlo (defaut : 42)"
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  CommandoQuant — Moteur VaR  (Mode Demo / Autonome)")
    print("=" * 60)

    if not os.path.exists(args.input):
        print(f"[ERREUR] Fichier d'entree introuvable : {args.input}")
        sys.exit(1)

    positions = load_positions(args.input)

    if not positions:
        print("[ERREUR] Aucune position valide trouvee dans le CSV.")
        sys.exit(1)

    print(f"\nPositions chargees ({len(positions)}) :")
    for p in positions:
        print(f"  {p['underlying']:<8} {p['option_type']:<5} "
              f"K={p['strike']:>7.2f}  S={p['spot']:>7.2f}  "
              f"Delta={p['delta']:>+.4f}")

    print("\n--- Calcul VaR ---")
    var_p              = var_parametrique(positions)
    var_mc, pnl_mc     = var_monte_carlo(positions, args.simulations, seed=args.seed)
    var_h, pnl_h       = var_historique(positions)

    print("\n--- Export du rapport ---")
    export_excel(positions, var_p, var_mc, pnl_mc, var_h, pnl_h, args.output)

    print("\n" + "=" * 60)
    print("  RESUME DES RESULTATS")
    print("=" * 60)
    print(f"  VaR Parametrique 95% : {var_p:>12,.2f} EUR")
    print(f"  VaR Monte Carlo  95% : {var_mc:>12,.2f} EUR")
    print(f"  VaR Historique   95% : {var_h:>12,.2f} EUR")
    print("=" * 60)
    print("  Termine.")


if __name__ == "__main__":
    main()
